"""Telegram bot that bridges messages to kimi-cli via ACP."""

from __future__ import annotations

import asyncio
import hashlib
import html
import json
import logging
import re
import shlex
import threading
from pathlib import Path
from typing import List, Optional

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.constants import ParseMode
from telegram.ext import (
    ApplicationBuilder,
    CallbackQueryHandler,
    CommandHandler,
    ContextTypes,
    MessageHandler,
    filters,
)

import config
from acp_client import AcpClient

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=getattr(logging, config.LOG_LEVEL.upper(), logging.INFO),
)
logger = logging.getLogger(__name__)


class BridgeState:
    """Holds the ACP client, session, and pending prompt state."""

    def __init__(self) -> None:
        self.acp: Optional[AcpClient] = None
        self.session_id: Optional[str] = None
        self.loop: Optional[asyncio.AbstractEventLoop] = None
        self.chat_id: Optional[int] = None
        self.bot = None

        self._buffer: list[str] = []
        self._buffer_lock = threading.Lock()
        self._prompt_done = asyncio.Event()
        self._busy = False
        self._current_stop_reason: Optional[str] = None
        self._pending_permissions: dict[int, threading.Event] = {}
        self._perm_lock = threading.Lock()

        # Context usage tracking (from usage_update notifications)
        self._context_used: int = 0
        self._context_size: int = 0
        self._context_lock = threading.Lock()

    # ------------------------------------------------------------------
    # Notification handler (runs on ACP reader thread)
    # ------------------------------------------------------------------

    def _on_permission_request(self, info: dict, event: threading.Event) -> None:
        """Called from ACP reader thread when agent requests permission."""
        req_id = info["request_id"]
        title = info.get("title", "Bilinmeyen işlem")
        with self._perm_lock:
            self._pending_permissions[req_id] = event

        if self.loop is not None and self.chat_id is not None:
            keyboard = [
                [InlineKeyboardButton("✅ Onayla", callback_data=f"perm:{req_id}:approve")],
                [InlineKeyboardButton("❌ Reddet", callback_data=f"perm:{req_id}:reject")],
            ]
            reply_markup = InlineKeyboardMarkup(keyboard)
            coro = self._send_permission_message(title, reply_markup)
            asyncio.run_coroutine_threadsafe(coro, self.loop)

    async def _send_permission_message(self, title: str, reply_markup) -> None:
        """Send permission request to Telegram chat."""
        try:
            if self.bot is None:
                return
            text = f"🔒 **İzin İsteği**\n\nKimi şu işlemi yapmak istiyor:\n`{html.escape(title)}`\n\nOnaylıyor musunuz?"
            await self.bot.send_message(
                chat_id=self.chat_id,
                text=text,
                parse_mode=ParseMode.MARKDOWN,
                reply_markup=reply_markup,
            )
        except Exception as exc:
            logger.exception("Failed to send permission message: %s", exc)

    def _on_acp_notification(self, msg: dict) -> None:
        if msg.get("method") != "session/update":
            return
        upd = msg.get("params", {}).get("update", {})
        t = upd.get("sessionUpdate")

        if t == "agent_message_chunk":
            text = upd.get("content", {}).get("text", "")
            with self._buffer_lock:
                self._buffer.append(text)

        elif t == "thought_message_chunk" or t == "agent_thought_chunk":
            # Thinking stream — ignore for final output
            pass

        elif t == "tool_call":
            title = upd.get("title", "")
            with self._buffer_lock:
                self._buffer.append(f"\n🔧 *{html.escape(title)}*\n")

        elif t == "tool_call_update":
            status = upd.get("status", "")
            # Only show final tool output (completed/failed), not input generation chunks
            if status in ("completed", "failed"):
                for block in upd.get("content", []):
                    if isinstance(block, dict) and block.get("type") == "content":
                        text = block.get("content", {}).get("text", "")
                        if text:
                            with self._buffer_lock:
                                self._buffer.append(text)
                if status == "completed":
                    with self._buffer_lock:
                        self._buffer.append("\n✅ Tamamlandı\n")
                elif status == "failed":
                    with self._buffer_lock:
                        self._buffer.append("\n❌ Başarısız\n")

        elif t == "available_commands":
            # Ignore slash-command ads from ACP
            pass

        elif t == "usage_update":
            size = upd.get("size", 0)
            used = upd.get("used", 0)
            with self._context_lock:
                self._context_size = size
                self._context_used = used
            logger.debug("Context usage updated: %s / %s", used, size)

        else:
            logger.debug("Unhandled session/update: %s", t)

    # ------------------------------------------------------------------
    # Prompt flow
    # ------------------------------------------------------------------

    def _context_footer(self) -> str:
        """Return a short context usage footer by reading context.jsonl."""
        used, size = _get_context_usage(self.session_id)
        if used == 0 or size == 0:
            return ""
        pct = (used / size) * 100
        # Format like: 19.2k / 262.1k (7.3%)
        def _fmt(n: int) -> str:
            if n >= 1_000_000:
                return f"{n/1_000_000:.1f}M"
            if n >= 1_000:
                return f"{n/1_000:.1f}k"
            return str(n)
        return f"\n\n📊 Context: {_fmt(used)} / {_fmt(size)} ({pct:.1f}%)"

    async def run_prompt(self, text: str) -> str:
        """Send a prompt and collect the streamed response."""
        logger.info("run_prompt: acp_ready=%s busy=%s", self.acp is not None and self.acp.is_ready(), self._busy)
        if self.acp is None or not self.acp.is_ready():
            return "⚠️ Kimi çalışmıyor. Önce /start komutuyla başlatın."

        if self._busy:
            return (
                "⏳ Şu an başka bir istek işleniyor. "
                "Bekleyin veya /cancel ile iptal edin."
            )

        self._busy = True
        with self._buffer_lock:
            self._buffer.clear()
        self._prompt_done.clear()
        self._current_stop_reason = None

        def _worker():
            logger.info("Prompt worker started")
            try:
                timeout = config.PROMPT_TIMEOUT if config.PROMPT_TIMEOUT > 0 else None
                resp = self.acp.prompt(self.session_id, text, timeout=timeout)
                logger.info("Prompt worker finished, resp=%s", resp)
                if resp:
                    self._current_stop_reason = resp.get("result", {}).get("stopReason")
            except Exception as exc:
                logger.exception("Prompt worker error: %s", exc)
            finally:
                logger.info("Prompt worker setting done event")
                if self.loop is not None:
                    self.loop.call_soon_threadsafe(self._prompt_done.set)

        threading.Thread(target=_worker, daemon=True).start()

        try:
            await asyncio.wait_for(
                self._prompt_done.wait(),
                timeout=config.PROMPT_TIMEOUT if config.PROMPT_TIMEOUT > 0 else None,
            )
        except asyncio.TimeoutError:
            self._busy = False
            return "⏰ Zaman aşımı. İstek çok uzun sürdü. /cancel ile iptal edebilir veya tekrar deneyebilirsiniz."

        # Give the reader thread a moment to flush final notifications
        await asyncio.sleep(0.5)

        with self._buffer_lock:
            result = "".join(self._buffer)

        self._busy = False
        logger.info("run_prompt finished, result_len=%d", len(result))

        if not result:
            return "⚠️ Kimi yanıt döndürmedi."
        return result + self._context_footer()

    def cancel_current(self) -> None:
        """Cancel the active prompt (if any)."""
        if self.acp and self.session_id:
            self.acp.cancel(self.session_id)


# Global state
_state = BridgeState()


# ------------------------------------------------------------------
# Telegram handlers
# ------------------------------------------------------------------

async def start_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user = update.effective_user
    if user is None or user.id not in config.AUTHORIZED_USERS:
        await update.message.reply_text("🚫 Yetkisiz erişim.")
        return

    if _state.acp is not None and _state.acp.is_ready():
        await update.message.reply_text("✅ Kimi zaten çalışıyor.")
        return

    mode = (context.args[0] if context.args else "auto").lower()
    if mode not in ("auto", "local", "vps"):
        await update.message.reply_text(
            "❌ Geçersiz mod. Kullanım:\n"
            "/start — Otomatik (local açıksa local, kapalıysa VPS)\n"
            "/start local — Her zaman local bilgisayardaki kimi CLI\n"
            "/start vps — Her zaman VPS'teki kimi CLI"
        )
        return

    await update.message.reply_text("🚀 Kimi başlatılıyor...")
    _state.loop = asyncio.get_running_loop()
    started = False

    # Build ACP client based on mode
    if mode == "local":
        inner_cmd = f"exec {config.KIMI_CMD} {' '.join(shlex.quote(a) for a in config.KIMI_ARGS)}"
        ssh_cmd = (
            f"ssh -p {shlex.quote(str(config.LOCAL_SSH_PORT))} "
            f"-i {shlex.quote(config.LOCAL_SSH_KEY)} "
            f"-o ConnectTimeout=2 "
            f"-o StrictHostKeyChecking=accept-new "
            f"-tt "
            f"{shlex.quote(f'{config.LOCAL_SSH_USER}@{config.LOCAL_SSH_HOST}')} "
            f"bash -lc {shlex.quote(inner_cmd)}"
        )
        _state.acp = AcpClient(
            cmd="script",
            args=["-q", "-c", ssh_cmd, "/dev/null"],
            cwd=config.LOCAL_SSH_WORK_DIR,
            notification_handler=_state._on_acp_notification,
        )
    elif mode == "vps":
        _state.acp = AcpClient(
            cmd=config.KIMI_CMD,
            args=config.KIMI_ARGS,
            cwd=config.WORK_DIR,
            notification_handler=_state._on_acp_notification,
        )
    else:  # auto
        # Try local first (via reverse SSH tunnel)
        inner_cmd = f"exec {config.KIMI_CMD} {' '.join(shlex.quote(a) for a in config.KIMI_ARGS)}"
        ssh_cmd = (
            f"ssh -p {shlex.quote(str(config.LOCAL_SSH_PORT))} "
            f"-i {shlex.quote(config.LOCAL_SSH_KEY)} "
            f"-o ConnectTimeout=2 "
            f"-o StrictHostKeyChecking=accept-new "
            f"-tt "
            f"{shlex.quote(f'{config.LOCAL_SSH_USER}@{config.LOCAL_SSH_HOST}')} "
            f"bash -lc {shlex.quote(inner_cmd)}"
        )
        local_acp = AcpClient(
            cmd="script",
            args=["-q", "-c", ssh_cmd, "/dev/null"],
            cwd=config.LOCAL_SSH_WORK_DIR,
            notification_handler=_state._on_acp_notification,
        )
        await update.message.reply_text("🔄 Local bilgisayar kontrol ediliyor...")
        started = local_acp.start()
        if started:
            _state.acp = local_acp
            mode = "local"
        else:
            local_acp.close()
            error_detail = local_acp.last_start_error or "Bilinmeyen hata"
            logger.warning("Local mode failed: %s", error_detail)
            await update.message.reply_text(
                f"⚠️ Local bağlantı başarısız:\n`{html.escape(error_detail)}`\n\n"
                f"VPS moduna geçiliyor...",
                parse_mode=ParseMode.MARKDOWN,
            )
            _state.acp = AcpClient(
                cmd=config.KIMI_CMD,
                args=config.KIMI_ARGS,
                cwd=config.WORK_DIR,
                notification_handler=_state._on_acp_notification,
            )
            mode = "vps"

    if _state.acp is None:
        await update.message.reply_text("❌ Kimi başlatılamadı.")
        return

    # local/vps modda veya auto'da fallback durumunda start() çağr
    if not started:
        ok = _state.acp.start()
        if not ok:
            error_detail = _state.acp.last_start_error or "Bilinmeyen hata"
            logger.error("Kimi start failed: %s", error_detail)
            await update.message.reply_text(
                f"❌ Kimi başlatılamadı:\n`{html.escape(error_detail)}`",
                parse_mode=ParseMode.MARKDOWN,
            )
            _state.acp = None
            return

    sid = _state.acp.new_session(_state.acp.cwd or config.WORK_DIR)
    if sid is None:
        await update.message.reply_text("❌ Oturum oluşturulamadı.")
        _state.acp.close()
        _state.acp = None
        return

    _state.session_id = sid
    mode_icon = "🖥️" if mode == "local" else "☁️"
    mode_text = "Local" if mode == "local" else "VPS"
    await update.message.reply_text(
        f"{mode_icon} Kimi hazır! ({mode_text} mod)\n"
        f"Session: `{sid}`\n"
        f"Dizin: `{_state.acp.cwd or config.WORK_DIR}`\n\n"
        f"Mesaj yazarak sohbete başlayabilirsiniz.\n"
        f"/stop ile durdurabilir, /cancel ile aktif işlemi iptal edebilirsiniz.",
        parse_mode=ParseMode.MARKDOWN,
    )


async def stop_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user = update.effective_user
    if user is None or user.id not in config.AUTHORIZED_USERS:
        return

    if _state.acp is None:
        await update.message.reply_text("ℹ️ Kimi zaten durdurulmuş.")
        return

    _state.acp.close()
    _state.acp = None
    _state.session_id = None
    _state._busy = False
    await update.message.reply_text("🛑 Kimi durduruldu.")


async def cancel_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user = update.effective_user
    if user is None or user.id not in config.AUTHORIZED_USERS:
        return

    if not _state._busy:
        await update.message.reply_text("ℹ️ Şu an işlemde bir istek yok.")
        return

    _state.cancel_current()
    await update.message.reply_text("🛑 İptal isteği gönderildi.")


async def help_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user = update.effective_user
    if user is None or user.id not in config.AUTHORIZED_USERS:
        return

    text = (
        "*Telegram Kimi Bridge*\n\n"
        "*Komutlar:*\n"
        "/start — Otomatik mod (local açıksa local, kapalıysa VPS)\n"
        "/start local — Her zaman local bilgisayardaki kimi CLI\n"
        "/start vps — Her zaman VPS'teki kimi CLI\n"
        "/stop — Kimi'yi kapat\n"
        "/cancel — Şu anki işlemi iptal et\n"
        "/help — Bu mesaj\n\n"
        "*Kullanım:*\n"
        "Doğrudan mesaj yazarak Kimi ile sohbet edebilirsiniz.\n"
        "Tool çağrıları (shell, dosya yazma vb.) için size onay isteği gönderilir.\n"
        "Uzun cevaplar otomatik parçalanır.\n"
        "Excel dosyaları otomatik okunur.\n\n"
        "*Not:* Kimi Code CLI ACP protokolü slash komutlarını (/clear, /plan vb.) "
        "henüz desteklemiyor. Bu komutları doğal dilde isteyebilirsiniz."
    )
    await update.message.reply_text(text, parse_mode=ParseMode.MARKDOWN)


async def callback_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle inline keyboard callbacks for permission requests."""
    query = update.callback_query
    await query.answer()

    data = query.data or ""
    if not data.startswith("perm:"):
        return

    parts = data.split(":")
    if len(parts) != 3:
        return

    _, req_id_str, action = parts
    try:
        req_id = int(req_id_str)
    except ValueError:
        return

    with _state._perm_lock:
        event = _state._pending_permissions.pop(req_id, None)
    if event is None:
        await query.edit_message_text("⏰ Bu izin isteği süresi dolmuş.")
        return

    if _state.acp is None:
        return

    if action == "approve":
        _state.acp._send_permission_response(req_id, "approve")
        await query.edit_message_text("✅ İzin onaylandı.")
    else:
        _state.acp._send_permission_response(req_id, "reject")
        await query.edit_message_text("❌ İzin reddedildi.")


def _extract_vba_blocks(text: str) -> List[str]:
    """Extract VBA code blocks from markdown text."""
    # Match ```vba ... ``` or ```vb ... ``` blocks
    pattern = r"```(?:vba|vb)\n(.*?)\n```"
    matches = re.findall(pattern, text, re.DOTALL | re.IGNORECASE)
    return [m.strip() for m in matches if m.strip()]


def _get_context_usage(session_id: Optional[str]) -> tuple[int, int]:
    """Read context usage from Kimi's context.jsonl and config.toml.
    Returns (used_tokens, max_context_size).
    """
    if not session_id:
        return 0, 0

    # 1. Find context.jsonl
    work_dir = config.WORK_DIR or "/home/akn/local"
    path_md5 = hashlib.md5(work_dir.encode("utf-8")).hexdigest()
    sessions_dir = Path.home() / ".kimi" / "sessions" / path_md5
    context_file = sessions_dir / session_id / "context.jsonl"

    used = 0
    if context_file.exists():
        try:
            with open(context_file, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line:
                        continue
                    try:
                        obj = json.loads(line)
                        if obj.get("role") == "_usage":
                            tc = obj.get("token_count")
                            if isinstance(tc, int):
                                used = tc
                    except json.JSONDecodeError:
                        continue
        except Exception as exc:
            logger.debug("Failed to read context.jsonl: %s", exc)

    # 2. Find max_context_size from config.toml
    size = 0
    config_file = Path.home() / ".kimi" / "config.toml"
    if config_file.exists():
        try:
            import tomllib
            with open(config_file, "rb") as f:
                cfg = tomllib.load(f)
            default_model = cfg.get("default_model", "")
            models = cfg.get("models", {})
            model_key = default_model
            if model_key in models:
                size = models[model_key].get("max_context_size", 0)
            # Try without quotes if needed (tomllib may parse quoted keys)
            if size == 0:
                for k, v in models.items():
                    if k.strip('"') == model_key or k == model_key:
                        size = v.get("max_context_size", 0)
                        break
        except Exception as exc:
            logger.debug("Failed to read config.toml: %s", exc)

    # Fallback for kimi-for-coding (Kimi-k2.6)
    if size == 0:
        size = 262144

    return used, size


def _read_excel_bytes(content_bytes: bytes, filename: str) -> str:
    """Read Excel file bytes and return a text summary."""
    import io
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(content_bytes), data_only=True)
        parts = [f"[Excel Dosyası: {filename}]"]
        parts.append(f"Sheet sayısı: {len(wb.sheetnames)}")
        parts.append(f"Sheet'ler: {', '.join(wb.sheetnames)}")
        parts.append("")

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            parts.append(f"--- Sheet: {sheet_name} ---")
            parts.append(f"Boyut: {ws.max_row} satır x {ws.max_column} sütun")
            parts.append("")

            # Show header + first 20 rows
            row_count = 0
            for row in ws.iter_rows(values_only=True):
                parts.append(" | ".join(str(cell) if cell is not None else "" for cell in row))
                row_count += 1
                if row_count >= 20:
                    parts.append("... [Satırlar kısaltıldı]")
                    break
            parts.append("")

        # Try to extract VBA modules if present (.xlsm)
        if filename.lower().endswith(".xlsm"):
            try:
                wb_vba = load_workbook(io.BytesIO(content_bytes), keep_vba=True)
                if wb_vba.vba_project:
                    parts.append("[VBA Projesi mevcut]")
                    # Note: openpyxl cannot read VBA source code directly,
                    # but it preserves the vbaProject.bin file
                    parts.append("(VBA kodunu okumak için dosyayı manuel olarak inceleyin)")
            except Exception:
                pass

        return "\n".join(parts)
    except Exception as exc:
        logger.exception("Excel read error: %s", exc)
        return f"[Excel dosyası okunamadı: {filename} — {exc}]"


async def message_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user = update.effective_user
    if user is None or user.id not in config.AUTHORIZED_USERS:
        return

    text = update.message.text or update.message.caption or ""

    # If a document is attached, try to download and read its content
    if update.message.document:
        doc = update.message.document
        logger.info("Document received: %s (%s bytes)", doc.file_name, doc.file_size)
        if doc.file_size and doc.file_size <= 5 * 1024 * 1024:  # max 5MB
            try:
                file = await context.bot.get_file(doc.file_id)
                content_bytes = await file.download_as_bytearray()

                fname = doc.file_name or "unknown"
                if fname.lower().endswith((".xlsx", ".xlsm", ".xls")):
                    # Excel file
                    excel_text = _read_excel_bytes(bytes(content_bytes), fname)
                    text = f"{text}\n\n{excel_text}"
                else:
                    # Try text decode for non-Excel files
                    try:
                        content = content_bytes.decode('utf-8')
                        max_file_chars = 30000
                        if len(content) > max_file_chars:
                            content = content[:max_file_chars] + "\n\n... [Dosya çok uzun, kısaltıldı]"
                        text = f"{text}\n\n[Dosya: {fname}]\n```\n{content}\n```"
                    except UnicodeDecodeError:
                        text = f"{text}\n\n[Dosya: {fname} — binary dosya, içerik okunamadı]"
            except Exception as exc:
                logger.exception("Failed to download/read file: %s", exc)
                text = f"{text}\n\n[Dosya indirilemedi: {doc.file_name}]"
        else:
            text = f"{text}\n\n[Dosya çok büyük (>{5}MB): {doc.file_name}]"

    if not text.strip():
        await update.message.reply_text("📎 Dosya alındı ama açıklama (caption) bulunamadı. Lütfen dosyanın yanına bir mesaj yazın.")
        return

    logger.info("Received message from %s: %s", user.id, text[:200])

    # Show typing indicator
    await context.bot.send_chat_action(
        chat_id=update.effective_chat.id, action="typing"
    )

    result = await _state.run_prompt(text)
    logger.info("Prompt result length: %d", len(result))

    # Check if response contains VBA code blocks — send as file too
    vba_blocks = _extract_vba_blocks(result)

    # Telegram message limit
    if len(result) <= config.MAX_MESSAGE_LENGTH:
        await update.message.reply_text(result)
    else:
        # Split into chunks
        for i in range(0, len(result), config.MAX_MESSAGE_LENGTH):
            chunk = result[i : i + config.MAX_MESSAGE_LENGTH]
            await update.message.reply_text(chunk)
            await asyncio.sleep(0.3)

    # Send VBA code as .vba document if found
    if vba_blocks:
        vba_content = "\n\n' --- Next Module ---\n\n".join(vba_blocks)
        await context.bot.send_document(
            chat_id=update.effective_chat.id,
            document=vba_content.encode("utf-8"),
            filename="kimi_macro.vba",
            caption="📎 VBA makro kodu (.vba dosyası)",
        )


# ------------------------------------------------------------------
# Entry point
# ------------------------------------------------------------------

def main() -> None:
    if not config.TELEGRAM_TOKEN:
        raise RuntimeError("TELEGRAM_TOKEN eksik! .env dosyasını kontrol edin.")
    if not config.AUTHORIZED_USERS:
        raise RuntimeError("AUTHORIZED_USERS eksik! .env dosyasını kontrol edin.")

    app = (
        ApplicationBuilder()
        .token(config.TELEGRAM_TOKEN)
        .build()
    )

    app.add_handler(CommandHandler("start", start_handler))
    app.add_handler(CommandHandler("stop", stop_handler))
    app.add_handler(CommandHandler("cancel", cancel_handler))
    app.add_handler(CommandHandler("help", help_handler))
    app.add_handler(CallbackQueryHandler(callback_handler))
    app.add_handler(MessageHandler((filters.TEXT | filters.PHOTO | filters.Document.ALL) & ~filters.COMMAND, message_handler))

    logger.info("Bot starting... Authorized users: %s", config.AUTHORIZED_USERS)
    _state.bot = app.bot
    app.run_polling()


if __name__ == "__main__":
    main()
